import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
import pandas as pd

def format_excel(file, row_number):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for cell in sheet[row_number]:
            cell.fill = yellow_fill

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"An error occurred: {e}")
        return None

def main():
    st.set_page_config(page_title="Excel Formatter Bot", page_icon="📊")

    st.title("Excel Formatter Bot")

    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file)
            st.write("File preview:")
            st.dataframe(df.head())

            row_number = st.number_input("Enter the row number to highlight:", min_value=1, value=1, step=1)

            if st.button("Format Excel"):
                formatted_file = format_excel(uploaded_file, row_number)

                if formatted_file:
                    st.download_button(
                        label="Download formatted Excel file",
                        data=formatted_file,
                        file_name="formatted_excel.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        except Exception as e:
            st.error(f"Could not read the file: {e}")

    st.write("Note: This bot uses openpyxl and may not support all Excel features. For complex Excel files, results may vary.")

if __name__ == "__main__":
    main()